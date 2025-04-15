#!/usr/bin/env python
# coding: utf-8

# In[27]:


import pandas as pd
import pybaseball as pyb
import numpy as np

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import pandas as pd


# In[28]:


#input the excel file path that you wish to update
excel_file_path = 'MLB Algorithm_2025 Season.xlsx'
print("Reading:",excel_file_path)

# ### Read in the injury report, parse it, then add it to the excel sheet

# In[29]:
step = 'Fangraphs Injury Report'

''' Read in fangraphs injury report '''
print('Reading in  %s',step)
# URL of the injury report for the 2023 season
url = 'https://www.fangraphs.com/roster-resource/injury-report?timeframe=all&season=2025'

# Send a GET request to fetch the page content
response = requests.get(url)
#response.raise_for_status()  # Ensure the request was successful

# Parse the HTML content using BeautifulSoup
soup = BeautifulSoup(response.text, 'html.parser')


# In[30]:


import requests
from bs4 import BeautifulSoup
import json
from tqdm import tqdm

# Step 1: Load page
url = 'https://www.fangraphs.com/roster-resource/injury-report?timeframe=all&season=2025'
response = requests.get(url)
soup = BeautifulSoup(response.text, 'html.parser')

# Step 2: Find the script tag with the JSON data
script_tag = soup.find('script', id='__NEXT_DATA__')

# Step 3: Parse the JSON
json_raw = script_tag.string
data = json.loads(json_raw)
res = []
# Step 4: Navigate to the injury data
injury_data = None
queries = data['props']['pageProps']['dehydratedState']['queries']
for q in queries:
    if isinstance(q.get('state', {}).get('data'), list):
        injury_data = q['state']['data']
        break

# Step 5: Use or print the injury data
if injury_data:
    for player in tqdm(injury_data):  # Preview first 5 entries
        # print(player)
        res.append({
            'Player': player.get('playerName'),
            'Team': player.get('team'),
            'Pos': player.get('position'),
            'Injury': player.get('injurySurgery'),
            'Status': player.get('status'),
            'Est. Return': player.get('returndate')
        })
else:
    print("Injury data not found.")

df = pd.DataFrame(res)
df.to_csv("Fangraphs_Injury_Report.csv")


file_path = excel_file_path

with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:            
    df.to_excel(writer, sheet_name="Injury", index=False, header=False)        

print("Wrote %s",step)

# ### Read in the daily lineup for rotowire, parse it, add to the excel sheet

# In[31]:

step = 'Fangraphs Injury Report'


print("Reading in %s",step)
url = "https://www.rotowire.com/baseball/daily-lineups.php"
soup = BeautifulSoup(requests.get(url).content, "html.parser")

data_pitiching = []
data_batter = []
team_type = ''

for e in soup.select('.lineup__box ul li'):
    if team_type != e.parent.get('class')[-1]:
        order_count = 1
        team_type = e.parent.get('class')[-1]

    if e.get('class') and 'lineup__player-highlight' in e.get('class'):
        data_pitiching.append({
            'date': e.find_previous('main').get('data-gamedate'),
            'game_time': e.find_previous('div', attrs={'class':'lineup__time'}).get_text(strip=True),
            'pitcher_name':e.a.get_text(strip=True),
            'team':e.find_previous('div', attrs={'class':team_type}).next.strip(),
            'lineup_throws':e.span.get_text(strip=True)
        })
    elif e.get('class') and 'lineup__player' in e.get('class'):
        data_batter.append({
            'date': e.find_previous('main').get('data-gamedate'),
            'game_time': e.find_previous('div', attrs={'class':'lineup__time'}).get_text(strip=True),
            'batter_name':e.a.get_text(strip=True),
            'team':e.find_previous('div', attrs={'class':team_type}).next.strip(),
            'pos': e.div.get_text(strip=True),
            'batting_order':order_count,
            'lineup_bats':e.span.get_text(strip=True)
        })
        order_count+=1

df_pitching = pd.DataFrame(data_pitiching)
df_batter = pd.DataFrame(data_batter)


# In[32]:


mlb_teams = {
    "Rockies": "COL",
    "Red Sox": "BOS",
    "Reds": "CIN",
    "Royals": "KC",
    "Diamondbacks": "ARI",
    "Marlins": "MIA",
    "Twins": "MIN",
    "Yankees": "NYY",
    "Angels": "LAA",
    "Braves": "ATL",
    "Nationals": "WSH",
    "Cardinals": "STL",
    "Phillies": "PHI",
    "Pirates": "PIT",
    "Astros": "HOU",
    "Dodgers": "LAD",
    "Rangers": "TEX",
    "Blue Jays": "TOR",
    "Orioles": "BAL",
    "White Sox": "CWS",
    "Padres": "SD",
    "Tigers": "DET",
    "Mets": "NYM",
    "Guardians": "CLE",
    "Brewers": "MIL",
    "Cubs": "CHC",
    "Giants": "SF",
    "Mariners": "SEA",
    "Athletics": "ATH",
    "Rays": "TB",
}


# In[33]:


combined = df_pitching.merge(df_batter,on=['team','date'],suffixes=("","_x"))
combined = combined[combined.columns.drop(list(combined.filter(regex='_x')))]
combined['team_short'] = combined['team'].apply(lambda x:mlb_teams[x])


# In[34]:


# Now need to make it in the order that it is in the excel sheet

combined = combined[['team_short','pitcher_name','batter_name','batting_order']]

file_path = excel_file_path

# Open the existing workbook
wb = load_workbook(file_path)

# Select the "Lineups" sheet
ws = wb['Lineups']

# Iterate over the rows of the DataFrame
for index, row in combined.iterrows():
    # Assign values to columns O, P, Q, R (index starts at 2 to leave space for header)
    ws[f'O{index + 2}'] = row['team_short']
    ws[f'P{index + 2}'] = row['pitcher_name']
    ws[f'Q{index + 2}'] = row['batter_name']
    ws[f'R{index + 2}'] = row['batting_order']

wb.save(file_path)
print("Wrote  %s",step)


# ### Update the table with Streamers Data (projected Rest of Season) + pybaseball's fangraphs dataset 

# ### Pull in the streamers data and the fangraphs data 

# In[35]:


import requests
from bs4 import BeautifulSoup
import json
from tqdm import tqdm
import pandas as pd
import pybaseball as pyb


# ##### Getting Batting Data

# In[ ]:

step = 'Batter Stats'
print("Reading in %s",step)
# Step 1: Load page
url = 'https://www.fangraphs.com/projections?pos=all&stats=bat&type=steameru'
response = requests.get(url)
soup = BeautifulSoup(response.text, 'html.parser')

# Step 2: Find the script tag with the JSON data
script_tag = soup.find('script', id='__NEXT_DATA__')

# Step 3: Parse the JSON
json_raw = script_tag.string
data = json.loads(json_raw)
res = []
# Step 4: Navigate to the injury data
streamers = None
queries = data['props']['pageProps']['dehydratedState']['queries']
for q in queries:
    if isinstance(q.get('state', {}).get('data'), list):
        streamers = q['state']['data']
        break

# Step 5: process the player data
if streamers:
    for player in tqdm(streamers):  # Preview first 5 entries
        # print(player)
        res.append({
            'Player': player.get('PlayerName', 'NA'),
            'Team': player.get('Team', 'NA'),
            'Pos': player.get('minpos', 'NA'),  # using 'minpos' as positional name
            'Age': player.get('Age', 'NA'),     # does not exist in sample
            'G': player.get('G', 'NA'),
            'AB': player.get('AB', 'NA'),
            'R': player.get('R', 'NA'),
            'H': player.get('H', 'NA'),
            '2B': player.get('2B', 'NA'),
            '3B': player.get('3B', 'NA'),
            'HR': player.get('HR', 'NA'),
            'RBI': player.get('RBI', 'NA'),
            'SB': player.get('SB', 'NA'),
            'CS': player.get('CS', 'NA'),
            'BB': player.get('BB', 'NA'),
            'SO': player.get('SO', 'NA'),
            'SH': player.get('SH', 'NA'),
            'SF': player.get('SF', 'NA'),
            'HBP': player.get('HBP', 'NA'),
            'AVG': player.get('AVG', 'NA'),
            'OBP': player.get('OBP', 'NA'),
            'SLG': player.get('SLG', 'NA'),
            'OPS': player.get('OPS', 'NA'),
            'Year': player.get('Year', 2025)   # does not exist in sample
        })
        
else:
    print("Streamers data not found in Fangraphs.")

df = pd.DataFrame(res)
df.to_csv("Fangraphs_Streamers_Data.csv")

df_batter = pyb.batting_stats(2025,qual=None)
df_batter = df_batter[['Name','Team','Pos','Age','G','AB','R','H','2B','3B','HR','RBI','SB','CS','BB','SO','SH','SF','HBP','AVG','OBP','SLG','OPS','Season']]
df_batter = df_batter.rename(columns={'Name':"Player","Season":'Year'})


# Need to do a function that gets the row of the streamers data for a given player, add their matching stats 
def combine_streamers_and_season_data(streamers_df, season_df):
    combined_rows = []
    
    # Iterate through each row in the season stats
    for _, season_row in season_df.iterrows():
        player_name = season_row['Player']
        
        # Find matching player in streamers data
        matching_streamer = streamers_df[streamers_df['Player'] == player_name]
        
        if not matching_streamer.empty:
            streamer_row = matching_streamer.iloc[0]
            
            # Combine rows using your combine logic
            combined = season_row.copy()
            for col in season_row.index:
                if col in streamer_row.index and col not in ['Player', 'Team', 'Pos', 'Age', 'Year']:
                    try:
                        combined[col] += streamer_row[col]
                    except:
                        combined[col] = 'NA'  # fallback in case of issues
                        
            # Recalculate rate stats
            combined['AVG'] = combined['H'] / combined['AB'] # H/ AB
            combined['OBP'] = (combined['H'] + combined['BB']) / (combined['AB'] + combined['BB'] - combined['SF'] - combined['SH']) # (H + BB) / (AB + BB - SF - SH)
            tb_no_1b = (combined['2B']*2) + (combined['3B'] * 3) + (combined['HR'] * 4) # Total Bases w/o Singles
            singles = combined['H'] - (combined['2B'] + combined['3B']+ combined['HR'])
            tb = tb_no_1b + singles
            combined['SLG'] = tb / combined['AB'] # tb / ab
            combined['OPS'] = combined['OBP'] + combined['SLG']
            combined['Pos'] = streamers_df['Pos']
            
            combined_rows.append(combined)
        else:
            # No matching streamer data, keep season data only
            combined_rows.append(season_row)
    
    # Create combined DataFrame
    return pd.DataFrame(combined_rows)

# Usage:
combined_data = combine_streamers_and_season_data(df, df_batter)


file_path = excel_file_path

# Open the existing workbook
wb = load_workbook(file_path)
ws =  wb['Batters']

for index, row in combined_data.iterrows():
    row_num = index + 3  # Start on row 2 to leave space for header

    ws[f'E{row_num}'] = row['Player']
    ws[f'F{row_num}'] = row['Team']
    ws[f'G{row_num}'] = str(row['Pos'])
    ws[f'H{row_num}'] = row['Age']
    ws[f'I{row_num}'] = row['G']
    ws[f'J{row_num}'] = row['AB']
    ws[f'K{row_num}'] = row['R']
    ws[f'L{row_num}'] = row['H']
    ws[f'M{row_num}'] = row['2B']
    ws[f'N{row_num}'] = row['3B']
    ws[f'O{row_num}'] = row['HR']
    ws[f'P{row_num}'] = row['RBI']
    ws[f'Q{row_num}'] = row['SB']
    ws[f'R{row_num}'] = row['CS']
    ws[f'S{row_num}'] = row['BB']
    ws[f'T{row_num}'] = row['SO']
    ws[f'U{row_num}'] = row['SH']
    ws[f'V{row_num}'] = row['SF']
    ws[f'W{row_num}'] = row['HBP']
    ws[f'X{row_num}'] = row['AVG']
    ws[f'Y{row_num}'] = row['OBP']
    ws[f'Z{row_num}'] = row['SLG']
    ws[f'AA{row_num}'] = row['OPS']
    ws[f'AB{row_num}'] = row['Year']

wb.save(file_path)

print("Wrote %s", step)


# ##### Getting Piching Data

# In[ ]:

step = 'Pitching Stats'
print("Reading %s", step)

# Step 1: Load page
url = 'https://www.fangraphs.com/projections?type=steameru&stats=pit&pos=&team=0&players=0&lg=all&z=1744628169&sortcol=&sortdir=desc&pageitems=30&statgroup=dashboard&fantasypreset=dashboard'
response = requests.get(url)
soup = BeautifulSoup(response.text, 'html.parser')

# Step 2: Find the script tag with the JSON data
script_tag = soup.find('script', id='__NEXT_DATA__')

# Step 3: Parse the JSON
json_raw = script_tag.string
data = json.loads(json_raw)
res = []
# Step 4: Navigate to the injury data
streamers = None
queries = data['props']['pageProps']['dehydratedState']['queries']
for q in queries:
    if isinstance(q.get('state', {}).get('data'), list):
        streamers = q['state']['data']
        break

# Step 5: process the player data
if streamers:
    for player in tqdm(streamers):  # Preview first 5 entries
        # print(player)
        res.append({
            'Player': player.get('PlayerName', 'NA'),
            'Team': player.get('Team', 'NA'),
            'Age': 'NA',  # Not available
            'G': player.get('G', 'NA'),
            'GS': player.get('GS', 'NA'),
            'CG': 'NA',  # Not available
            'ShO': 'NA',  # Not available
            'IP': player.get('IP', 'NA'),
            'H': player.get('H', 'NA'),
            'ER': player.get('ER', 'NA'),
            'SO': player.get('SO', 'NA'),  # Note: 'SO' is K
            'BB': player.get('BB', 'NA'),
            'HR': player.get('HR', 'NA'),
            'W': player.get('W', 'NA'),
            'L': player.get('L', 'NA'),
            'SV': player.get('SV', 'NA'),
            'BS': player.get('BS', 'NA'),
            'HLD': player.get('HLD', 'NA'),
            'ERA': player.get('ERA', 'NA'),
            'WHIP': player.get('WHIP', 'NA'),
            'Year': 2025,
            'IP per GS': player.get('IP', 0) / player.get('GS', 1) if player.get('GS', 1) != 0 else 0
    })
        
else:
    print("Streamers data not found in Fangraphs.")

df = pd.DataFrame(res)
df.to_csv("Fangraphs_Streamers_Data.csv")

df_pitcher = pyb.pitching_stats(2025, qual=None)
df_pitcher = df_pitcher[['Name','Team','Age','G','GS','CG','ShO','IP','H','ER','SO','BB','HR','W','L','SV','BS','HLD','ERA','WHIP','Season']]
df_pitcher['IP per GS'] = df_pitcher['IP'] / df_pitcher['GS'].replace(0, pd.NA)
df_pitcher = df_pitcher.rename(columns={'Name':"Player","Season":'Year'})


# Need to do a function that gets the row of the streamers data for a given player, add their matching stats 
def combine_pitcher_data(streamers_df, season_df):
    combined_rows = []

    for idx, season_row in season_df.iterrows():
        player_name = season_row['Player']
        match = streamers_df[streamers_df['Player'] == player_name]

        if not match.empty:
            stream_row = match.iloc[0]
            combined = season_row.copy()

            for col in season_row.index:
                if col in stream_row.index and col not in ['Player', 'Team', 'Age', 'Year', 'IP per GS']:
                    try:
                        combined[col] += stream_row[col]
                    except:
                        combined[col] = 'NA'

            # Recalculate derived stat
            combined['IP per GS'] = combined['IP'] / combined['GS'] if combined['GS'] else 0
            combined['ERA'] = (combined['ER'] / combined['IP']) * 9
            combined['WHIP'] = (combined['BB'] + combined['H']) / combined['IP']

            combined_rows.append(combined)
        else:
            combined_rows.append(season_row)

    return pd.DataFrame(combined_rows)

# Usage:
combined = combine_pitcher_data(df, df_pitcher)


file_path = excel_file_path

wb = load_workbook(file_path)

ws = wb['Pitchers']

# Iterate over the rows of the DataFrame

for index, row in combined.iterrows():
    row_num = index + 2  # Start on row 2 to leave space for header
    
    ws[f'B{row_num}'] = row['Player']
    ws[f'C{row_num}'] = row['Team']
    ws[f'D{row_num}'] = row['Age']
    ws[f'E{row_num}'] = row['G']
    ws[f'F{row_num}'] = row['GS']
    ws[f'G{row_num}'] = row['CG']
    ws[f'H{row_num}'] = row['ShO']
    ws[f'I{row_num}'] = row['IP']
    ws[f'J{row_num}'] = row['H']
    ws[f'K{row_num}'] = row['ER']
    ws[f'L{row_num}'] = row['SO']
    ws[f'M{row_num}'] = row['BB']
    ws[f'N{row_num}'] = row['HR']
    ws[f'O{row_num}'] = row['W']
    ws[f'P{row_num}'] = row['L']
    ws[f'Q{row_num}'] = row['SV']
    ws[f'R{row_num}'] = row['BS']
    ws[f'S{row_num}'] = row['HLD']
    ws[f'T{row_num}'] = row['ERA']
    ws[f'U{row_num}'] = row['WHIP']
    ws[f'V{row_num}'] = row['Year']
    ws[f'W{row_num}'] = row['IP per GS']

wb.save(file_path)       


# In[43]:




file_path = excel_file_path

wb = load_workbook(file_path)

ws = wb['Pitchers']

# Iterate over the rows of the DataFrame

for index, row in combined.iterrows():
    row_num = index + 2  # Start on row 2 to leave space for header
    
    ws[f'B{row_num}'] = row['Player']
    ws[f'C{row_num}'] = row['Team']
    ws[f'D{row_num}'] = row['Age']
    ws[f'E{row_num}'] = row['G']
    ws[f'F{row_num}'] = row['GS']
    ws[f'G{row_num}'] = row['CG']
    ws[f'H{row_num}'] = row['ShO']
    ws[f'I{row_num}'] = row['IP']
    ws[f'J{row_num}'] = row['H']
    ws[f'K{row_num}'] = row['ER']
    ws[f'L{row_num}'] = row['SO']
    ws[f'M{row_num}'] = row['BB']
    ws[f'N{row_num}'] = row['HR']
    ws[f'O{row_num}'] = row['W']
    ws[f'P{row_num}'] = row['L']
    ws[f'Q{row_num}'] = row['SV']
    ws[f'R{row_num}'] = row['BS']
    ws[f'S{row_num}'] = row['HLD']
    ws[f'T{row_num}'] = row['ERA']
    ws[f'U{row_num}'] = row['WHIP']
    ws[f'V{row_num}'] = row['Year']
    ws[f'W{row_num}'] = row['IP per GS']

wb.save(file_path)   
  
print("Wrote %s", step)


# In[ ]:




